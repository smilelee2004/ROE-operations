import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from tqdm import tqdm
import threading
import tkinter as tk
from tkinter import messagebox
from tkinter import ttk  # 導入 ttk 模組

def read_xls_column_to_list(file_path):
    # 讀取 xlsm 檔案
    wb = load_workbook(file_path, read_only=True)
    sheet = wb.active
    
    # 取得第一欄 (Column 1) 的資料並轉換成 list
    column_1_list = [row[0] for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)]
    
    return column_1_list

def process_files(base_path, file_list, output_file, progress_var, cancel_event, root):
    all_data = []
    count = 0
    headers_written = False

    for file_name in tqdm(file_list, desc="Processing files", unit="file"):
        if cancel_event.is_set():
            print("Processing cancelled")
            break

        # 直接使用 file_name 來生成 file_path
        file_path = os.path.join(base_path, f"{file_name}.xlsm")
        
        if os.path.exists(file_path):
            print(f"Processing file: {file_path}")
            try:
                # 讀取對應的 xlsm 檔案
                wb = load_workbook(file_path, data_only=True, read_only=False)
                sheet = wb["美股"]  # 讀取名叫 "美股" 的工作表
                
                # 取得 O10 到 P15 區間的最大值與最小值
                o10_p15 = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=10, max_row=15, min_col=15, max_col=16)]
                max_value = max(max(row) for row in o10_p15)  # 取得最大值
                min_value = min(min(row) for row in o10_p15)  # 取得最小值
                print(f"ROE {sheet.cell(row=13, column=22).value}")
                # 取得特定欄位的資料
                data = {
                    '代號': file_name,  # 加入 file_name
                    'ROE': sheet.cell(row=13, column=22).value,  # V13
                    '手調貴': max_value,
                    '手調淑': min_value,
                    '貴價': sheet.cell(row=7, column=11).value,    # K7
                    '淑價': sheet.cell(row=5, column=11).value,    # K5
                    '現價': sheet.cell(row=3, column=11).value,    # K3
                    '預期報酬': sheet.cell(row=4, column=11).value,    # K4
                    '財報': sheet.cell(row=24, column=1).value,   # A24
                    '檔案路徑': f'=HYPERLINK("{file_path}", "點我開啟檔案")'    # 加入 file_path
                }
            except Exception as e:
                print(f"Failed to process file: {file_path}, error: {e}")
                # 如果讀取失敗，只填入 file_name，其他欄位保持空白
                data = {
                    '代號': file_name,
                    'ROE': None,
                    '手調貴': None,
                    '手調淑': None,
                    '貴價': None,
                    '淑價': None,
                    '現價': None,
                    '預期報酬': None,
                    '財報': None,
                    '檔案路徑': f'=HYPERLINK("{file_path}", "點我開啟檔案")'
                }
        else:
            print(f"File not found: {file_path}")
            data = {
                '代號': file_name,
                'ROE': None,
                '手調貴': None,
                '手調淑': None,
                '貴價': None,
                '淑價': None,
                '現價': None,
                '預期報酬': None,
                '財報': None,
                '檔案路徑': f'=HYPERLINK("{file_path}", "點我開啟檔案")'
            }
        all_data.append(data)
        count += 1
        
        # 更新進度條
        progress_var.set(count)
        root.update_idletasks()
        
        # 每 10 筆資料寫檔一次
        if count % 10 == 0:
            combined_data = pd.DataFrame(all_data)
            if not headers_written:
                combined_data.to_excel(output_file, sheet_name='美股', index=False, engine='openpyxl')
                headers_written = True
            else:
                with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    combined_data.to_excel(writer, sheet_name='美股', index=False, header=False, startrow=writer.sheets['美股'].max_row)
            all_data = []  # 清空 all_data

    # 寫入剩餘的資料
    if all_data:
        combined_data = pd.DataFrame(all_data)
        if not headers_written:
            combined_data.to_excel(output_file, sheet_name='美股', index=False, engine='openpyxl')
        else:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                combined_data.to_excel(writer, sheet_name='美股', index=False, header=False, startrow=writer.sheets['美股'].max_row)

def main():
    base_path = r"D:\work\me\what\company\system\資訊處理循環\tools\盈再表"  # 替換成實際的檔案路徑
    a_file_path = r"D:\work\me\what\company\system\資訊處理循環\tools\盈再表\a.xlsx"
    
    # 生成以當日日期為檔名的 xlsx 檔案
    today_date = datetime.now().strftime("%Y%m%d")
    output_file = os.path.join(base_path, f'{today_date}.xlsx')
    
    # 呼叫函數，讀取 a_file_path 檔案並將第一欄資料存入 list
    file_list = read_xls_column_to_list(a_file_path)
    
    # 創建進度條和取消按鈕
    root = tk.Tk()
    root.title("Processing Files")
    
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=len(file_list))  # 使用 ttk.Progressbar
    progress_bar.pack(fill=tk.X, expand=1, padx=10, pady=10)
    
    cancel_event = threading.Event()
    
    def cancel():
        cancel_event.set()
        messagebox.showinfo("Cancelled", "Processing has been cancelled.")
    
    cancel_button = tk.Button(root, text="Cancel", command=cancel)
    cancel_button.pack(pady=10)
    
    def run_processing():
        process_files(base_path, file_list, output_file, progress_var, cancel_event, root)
        root.quit()
    
    threading.Thread(target=run_processing).start()
    root.mainloop()

# 檢查是否直接執行此檔案
if __name__ == "__main__":
    main()