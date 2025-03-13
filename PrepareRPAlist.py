import pandas as pd
from openpyxl import load_workbook, Workbook
import os
import threading
import tkinter as tk
from tkinter import ttk, messagebox
import time

cancel_flag = False

def create_empty_excel(file_path):
    wb = Workbook()
    wb.save(file_path)

def copy_column_a_to_b(a_file, b_file):
    if not os.path.exists(b_file):
        create_empty_excel(b_file)
    
    df_a = pd.read_excel(a_file, usecols=[0], header=None)
    
    with pd.ExcelWriter(b_file, engine='openpyxl', mode='w') as writer:
        if not writer.book.sheetnames:
            writer.book.create_sheet("Sheet1")
        sheet_name = writer.book.sheetnames[0]
        df_a.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startcol=0)

def copy_b_to_k(b_file, progress_var, root):
    global cancel_flag
    wb = load_workbook(b_file)
    ws = wb[wb.sheetnames[0]]
    ws2 = wb.create_sheet("Sheet2")
    all_data = []
    historicalPeiord = int(time.time())  # 從1970/1/1到今日的總秒數

    data = {
        'StockID': "StockID",
        'QuarterlyIncomeStatement': f"QuarterlyIncomeStatement",
        'QuarterlyBalanceSheet': f"QuarterlyBalanceSheet",
        'AnnualIncomeStatement': f"AnnualIncomeStatement",
        'AnnualBalanceSheet': f"AnnualBalanceSheet",
        'CompanyNameProfile': f"CompanyNameProfile",
        'AnnualCashflowStatement': f"AnnualCashflowStatement",
        'HistoricalStockPrice': f"HistoricalStockPrice",
    }
    all_data.append(data)
    max_row = ws.max_row
    
    for row in range(2, max_row + 1):
        if cancel_flag:
            print("Operation cancelled.")
            break
        testString = ws.cell(row, 1).value.replace('.', '-')
        data = {
            'StockID': ws.cell(row, 1).value,
            'QuarterlyIncomeStatement': f"https://www.marketwatch.com/investing/Stock/{ws.cell(row, 1).value}/financials/income/quarter",
            'QuarterlyBalanceSheet': f"https://www.marketwatch.com/investing/Stock/{ws.cell(row, 1).value}/financials/balance-sheet/quarter",
            'AnnualIncomeStatement': f"https://www.marketwatch.com/investing/Stock/{ws.cell(row, 1).value}/financials",
            'AnnualBalanceSheet': f"https://www.marketwatch.com/investing/Stock/{ws.cell(row, 1).value}/financials/balance-sheet",
            'CompanyNameProfile': f"https://www.marketwatch.com/investing/stock/{ws.cell(row, 1).value}/company-profile",
            'AnnualCashflowStatement': f"https://www.marketwatch.com/investing/Stock/{ws.cell(row, 1).value}/financials/cash-flow",
            'HistoricalStockPrice': f"https://finance.yahoo.com/quote/{testString}/history?period1=573436800&period2={historicalPeiord}&interval=1mo&filter=history&frequency=1mo",
        }
        all_data.append(data)
        progress_var.set((row - 1) / (max_row - 1) * 100)
        root.update_idletasks()

    for data in all_data:
        ws2.append(list(data.values()))

    wb.remove(ws)
    wb.save(b_file)

def cancel_operation():
    global cancel_flag
    cancel_flag = True

def main():
    global cancel_flag
    cancel_flag = False

    a_file = r"D:\work\me\what\company\system\資訊處理循環\tools\盈再表\a.xlsx"
    b_file = r"D:\work\me\what\company\system\資訊處理循環\tools\RPA\RPA下載清單.xlsx"
    
    root = tk.Tk()
    root.title("Progress")
    root.geometry("300x100")

    progress_label = tk.Label(root, text="Preparing RPAList...")
    progress_label.pack(pady=10)

    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
    progress_bar.pack(pady=10)

    cancel_button = tk.Button(root, text="Cancel", command=cancel_operation)
    cancel_button.pack(pady=10)

    threading.Thread(target=lambda: [copy_column_a_to_b(a_file, b_file), copy_b_to_k(b_file, progress_var, root), root.quit()]).start()
    root.mainloop()

if __name__ == "__main__":
    main()